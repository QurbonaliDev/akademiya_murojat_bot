# export.py
# Excel va CSV export funksiyalari (rangli dizayn bilan)

import pandas as pd
import csv
import logging
from datetime import datetime
from database import get_all_complaints
from config import DATABASE_NAME
import sqlite3
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def export_to_excel():
    """Murojaatlarni Excel fayliga eksport qilish (rangli dizayn bilan)"""
    try:
        conn = sqlite3.connect(DATABASE_NAME)

        df = pd.read_sql_query('''
            SELECT 
                id as "ID",
                direction as "Yo'nalish",
                course as "Kurs", 
                complaint_type as "Murojaat turi",
                subject_name as "Fan nomi",
                teacher_name as "O'qituvchi",
                message as "Xabar",
                created_at as "Sana"
            FROM complaints 
            ORDER BY created_at DESC
        ''', conn)

        conn.close()

        filename = f"murojaatlar_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        df.to_excel(filename, index=False, engine='openpyxl')

        # === Excel faylni ochamiz va dizayn beramiz ===
        wb = load_workbook(filename)
        ws = wb.active

        # --- Stil sozlamalari ---
        header_fill = PatternFill("solid", fgColor="2E7D32")  # To‘q yashil
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # --- Har bir katakni sozlash ---
        for r_idx, row in enumerate(ws.iter_rows(), start=1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                if r_idx == 1:  # sarlavha qatori
                    cell.fill = header_fill
                    cell.font = header_font

        # --- Ustun kengligini avtomatik belgilash ---
        max_col_width = 40
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_length:
                    max_length = len(val)
            ws.column_dimensions[col_letter].width = min(max_length + 2, max_col_width)

        # --- Filtr qo‘shish ---
        ws.auto_filter.ref = ws.dimensions

        wb.save(filename)

        logger.info(f"Excel fayl muvaffaqiyatli rangli tarzda yaratildi: {filename}")
        return filename

    except Exception as e:
        logger.error(f"Excel export xatosi: {e}")
        return None
