# export.py
# Excel export funksiyalari (Professional ko'k dizayn bilan)

import logging
import sqlite3
from datetime import datetime
import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font

from config.config import DATABASE_NAME

logger = logging.getLogger(__name__)

# --- Stil konstantalari (Professional ko'k tema) ---
BLUE_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")  # Professional to'q ko'k
WHITE_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_styling(ws):
    """Excel varag'iga professional dizayn berish"""
    # Har bir katakni sozlash
    for r_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN

            if r_idx == 1:  # Sarlavha qatori
                cell.fill = BLUE_HEADER_FILL
                cell.font = WHITE_FONT
                cell.alignment = CENTER_ALIGN

    # Ustun kengliklarini avtomatik sozlash
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 4, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Filtr qo'shish
    ws.auto_filter.ref = ws.dimensions

def export_to_excel():
    """Murojaatlarni Excel fayliga eksport qilish"""
    try:
        conn = sqlite3.connect(DATABASE_NAME)
        
        # Murojaatlarni o'qish
        df = pd.read_sql_query('''
            SELECT 
                uid as "ID",
                created_at as "Sana",
                course as "Kurs",
                faculty as "Fakultet",
                direction as "Yo'nalish",
                subject_name as "Fan",
                teacher_name as "O'qituvchi",
                complaint_type as "Turi",
                message as "Xabar",
                status as "Status"
            FROM complaints 
            ORDER BY created_at DESC
        ''', conn)
        conn.close()

        filename = f"murojaatlar_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        # Excelga saqlash
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Raw_Data')
            
            ws = writer.sheets['Raw_Data']
            apply_styling(ws)

        logger.info(f"Murojaatlar eksport qilindi: {filename}")
        return filename

    except Exception as e:
        logger.error(f"Complaint export xatosi: {e}")
        return None

def export_to_excel_for_lesson_ratings():
    """Dars baholashlarini professional eksport qilish"""
    try:
        conn = sqlite3.connect(DATABASE_NAME)
        
        # Ma'lumotlarni o'qish (Yangi 1-qatorli strukturadan)
        df = pd.read_sql_query('''
            SELECT 
                uid as "ID",
                created_at as "Sana",
                course as "Kurs",
                faculty as "Fakultet",
                direction as "Yo'nalish",
                subject_name as "Fan",
                teacher_name as "O'qituvchi",
                q1 as "Savol 1",
                q2 as "Savol 2",
                q3 as "Savol 3",
                q4 as "Savol 4",
                q5 as "Savol 5",
                q6 as "Savol 6",
                total_score as "Umumiy baho",
                status as "Status"
            FROM lesson_ratings
            ORDER BY created_at DESC
        ''', conn)
        conn.close()

        if df.empty:
            return None

        # Bo'sh q7-q10 ustunlarini rasmda yo'qligi uchun chiqarmaymiz
        
        filename = f"baholashlar_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Raw_Data')
            ws = writer.sheets['Raw_Data']
            apply_styling(ws)

        logger.info(f"Baholashlar eksport qilindi: {filename}")
        return filename

    except Exception as e:
        logger.error(f"Rating export xatosi: {e}")
        return None
